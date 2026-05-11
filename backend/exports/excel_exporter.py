"""
Excel & CSV Exporter — generates final output files.

Output format matches the professor's requirements:
- Excel Sheet 1: Summary (Vehicle Type | Count)
- Excel Sheet 2: Detailed crossing log (track_id | vehicle_type | timestamp)
- CSV: Simple summary format
"""

import logging
import csv
from datetime import datetime
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# Project root
PROJECT_ROOT = Path(__file__).parent.parent.parent
EXCEL_OUTPUT_DIR = PROJECT_ROOT / "outputs" / "excel"


def _ensure_output_dir():
    """Create output directory if it doesn't exist."""
    EXCEL_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)


def export_to_excel(
    counts: dict[str, int],
    video_name: str,
    crossing_log: Optional[list[dict]] = None,
    video_metadata: Optional[dict] = None,
    output_path: Optional[str] = None,
) -> str:
    """
    Export counting results to a styled Excel file.

    Args:
        counts: Vehicle type → count mapping.
        video_name: Name of the processed video.
        crossing_log: Optional detailed crossing records.
        video_metadata: Optional video info (fps, duration, etc.).
        output_path: Custom output path. If None, auto-generates in outputs/excel/.

    Returns:
        Path to the generated Excel file.
    """
    _ensure_output_dir()

    if output_path is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_name = Path(video_name).stem.replace(" ", "_")
        output_path = str(EXCEL_OUTPUT_DIR / f"{safe_name}_{timestamp}.xlsx")

    wb = Workbook()

    # ── Sheet 1: Summary ──
    ws_summary = wb.active
    ws_summary.title = "Vehicle Counts"

    # Styles
    header_font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    data_font = Font(name="Calibri", size=11)
    total_font = Font(name="Calibri", bold=True, size=12, color="2F5496")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Title
    ws_summary["A1"] = f"Traffic Count Report — {video_name}"
    ws_summary["A1"].font = Font(name="Calibri", bold=True, size=14, color="2F5496")
    ws_summary.merge_cells("A1:B1")

    # Timestamp
    ws_summary["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws_summary["A2"].font = Font(name="Calibri", size=9, color="808080")

    # Video metadata
    row = 4
    if video_metadata:
        ws_summary[f"A{row}"] = "Video Information"
        ws_summary[f"A{row}"].font = Font(name="Calibri", bold=True, size=11)
        row += 1
        for key, value in video_metadata.items():
            ws_summary[f"A{row}"] = key.replace("_", " ").title()
            ws_summary[f"B{row}"] = str(value)
            ws_summary[f"A{row}"].font = data_font
            ws_summary[f"B{row}"].font = data_font
            row += 1
        row += 1

    # Headers
    ws_summary[f"A{row}"] = "Vehicle Type"
    ws_summary[f"B{row}"] = "Count"
    for col in ["A", "B"]:
        cell = ws_summary[f"{col}{row}"]
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
    row += 1

    # Data rows
    for vehicle_type, count in counts.items():
        if count > 0 or vehicle_type != "Unknown":
            ws_summary[f"A{row}"] = vehicle_type
            ws_summary[f"B{row}"] = count
            ws_summary[f"A{row}"].font = data_font
            ws_summary[f"B{row}"].font = data_font
            ws_summary[f"B{row}"].alignment = Alignment(horizontal="center")
            for col in ["A", "B"]:
                ws_summary[f"{col}{row}"].border = thin_border
            row += 1

    # Total row
    total = sum(counts.values())
    ws_summary[f"A{row}"] = "TOTAL"
    ws_summary[f"B{row}"] = total
    ws_summary[f"A{row}"].font = total_font
    ws_summary[f"B{row}"].font = total_font
    ws_summary[f"B{row}"].alignment = Alignment(horizontal="center")
    for col in ["A", "B"]:
        ws_summary[f"{col}{row}"].border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="double"),
            bottom=Side(style="double"),
        )

    # Column widths
    ws_summary.column_dimensions["A"].width = 25
    ws_summary.column_dimensions["B"].width = 15

    # ── Sheet 2: Detailed Crossing Log ──
    if crossing_log:
        ws_detail = wb.create_sheet("Crossing Details")

        detail_headers = ["Track ID", "Vehicle Type", "Timestamp (s)", "Direction", "Position X", "Position Y"]
        for col_idx, header in enumerate(detail_headers, 1):
            cell = ws_detail.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        for row_idx, record in enumerate(crossing_log, 2):
            ws_detail.cell(row=row_idx, column=1, value=record.get("track_id", "")).font = data_font
            ws_detail.cell(row=row_idx, column=2, value=record.get("vehicle_type", "")).font = data_font
            ws_detail.cell(row=row_idx, column=3, value=round(record.get("timestamp", 0), 2)).font = data_font
            ws_detail.cell(row=row_idx, column=4, value=record.get("direction", "")).font = data_font

            pos = record.get("position", (0, 0))
            ws_detail.cell(row=row_idx, column=5, value=round(pos[0], 1)).font = data_font
            ws_detail.cell(row=row_idx, column=6, value=round(pos[1], 1)).font = data_font

            for col_idx in range(1, 7):
                ws_detail.cell(row=row_idx, column=col_idx).border = thin_border
                ws_detail.cell(row=row_idx, column=col_idx).alignment = Alignment(horizontal="center")

        # Auto-width
        for col_idx in range(1, 7):
            ws_detail.column_dimensions[get_column_letter(col_idx)].width = 18

    # Save
    wb.save(output_path)
    logger.info(f"Excel exported: {output_path}")
    return output_path


def export_to_csv(
    counts: dict[str, int],
    video_name: str,
    output_path: Optional[str] = None,
) -> str:
    """
    Export counting results to CSV.

    Args:
        counts: Vehicle type → count mapping.
        video_name: Name of the processed video.
        output_path: Custom output path.

    Returns:
        Path to the generated CSV file.
    """
    _ensure_output_dir()

    if output_path is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_name = Path(video_name).stem.replace(" ", "_")
        output_path = str(EXCEL_OUTPUT_DIR / f"{safe_name}_{timestamp}.csv")

    with open(output_path, "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["Vehicle Type", "Count"])
        for vehicle_type, count in counts.items():
            writer.writerow([vehicle_type, count])
        writer.writerow(["TOTAL", sum(counts.values())])

    logger.info(f"CSV exported: {output_path}")
    return output_path
